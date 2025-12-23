import os
import pandas as pd
import qrcode
from flask import Flask, render_template, request, send_file
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

# ---------------- FLASK APP ----------------
app = Flask(__name__)

# ---------------- PATHS ----------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

UPLOAD_FOLDER = os.path.join(BASE_DIR, "uploads")
OUTPUT_FOLDER = os.path.join(BASE_DIR, "outputs")
QR_FOLDER = os.path.join(BASE_DIR, "qr_codes")

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)
os.makedirs(QR_FOLDER, exist_ok=True)

# ---------------- QR CONFIG ----------------
QR_MM = 9
PIXELS_PER_MM = 96 / 25.4
QR_PX = int(QR_MM * PIXELS_PER_MM)
ROW_HEIGHT = 30
QR_COL_WIDTH = 18
PATH_COL_WIDTH = 95
# ------------------------------------------


def normalize(text):
    return (
        str(text)
        .replace("\xa0", " ")
        .replace("\u200b", "")
        .strip()
        .lower()
    )


def generate_qr(data, path):
    qr = qrcode.QRCode(
        version=None,
        error_correction=qrcode.constants.ERROR_CORRECT_M,
        box_size=10,
        border=1,
    )
    qr.add_data(data)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    img.save(path)


def build_qr_text(row, selected_cols):
    """
    IMPORTANT RULE:
    - Values are used EXACTLY AS READ (NO formatting change)
    - Empty â†’ |
    - Separator â†’ space
    """
    parts = []
    for col in selected_cols:
        val = row.get(col, "")
        if pd.isna(val) or str(val).strip() == "":
            parts.append("|")
        else:
            # KEEP ORIGINAL VALUE AS STRING
            parts.append(str(val))
    return " ".join(parts)


# ---------------- ROUTES ----------------

@app.route("/", methods=["GET", "POST"])
def upload():
    if request.method == "POST":
        file = request.files["excel"]
        file_path = os.path.join(UPLOAD_FOLDER, file.filename)
        file.save(file_path)

        # READ AS STRING â†’ prevents 95,000 â†’ 95000 or 1.00 â†’ 1
        df = pd.read_excel(file_path, dtype=str)
        columns = list(df.columns)

        return render_template(
            "columns.html",
            columns=columns,
            file_path=file_path
        )

    return render_template("upload.html")


@app.route("/generate", methods=["POST"])
def generate():
    file_path = request.form["file_path"]
    selected_cols = request.form.getlist("columns")

    # READ EVERYTHING AS STRING
    df = pd.read_excel(file_path, dtype=str)
    df_norm = df.copy()
    df_norm.columns = [normalize(c) for c in df_norm.columns]
    selected_norm = [normalize(c) for c in selected_cols]

    df["QR_Image_Path"] = ""
    df["QR"] = ""

    for i, row in df_norm.iterrows():
        qr_text = build_qr_text(row, selected_norm)

        img_name = f"qr_{i+1}.png"
        img_path = os.path.join(QR_FOLDER, img_name)

        # ðŸ”‘ PUBLIC URL (NOT SERVER PATH)
        public_qr_url = request.host_url.rstrip("/") + "/qr/" + img_name

        generate_qr(public_qr_url, img_path)

        df.at[i, "QR_Image_Path"] = img_path
        df.at[i, "QR"] = qr_text

    output_path = os.path.join(OUTPUT_FOLDER, "output_with_qr.xlsx")
    df.to_excel(output_path, index=False)

    # INSERT QR IMAGES
    wb = load_workbook(output_path)
    ws = wb.active

    headers = {normalize(c.value): i + 1 for i, c in enumerate(ws[1])}
    qr_col = headers["qr"]
    path_col = headers["qr_image_path"]

    for r in range(2, ws.max_row + 1):
        img_path = ws.cell(r, path_col).value
        if not img_path or not os.path.exists(img_path):
            continue

        img = Image(img_path)
        img.width = QR_PX
        img.height = QR_PX

        ws.add_image(img, ws.cell(r, qr_col).coordinate)
        ws.row_dimensions[r].height = ROW_HEIGHT
        ws.cell(r, qr_col).alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    ws.column_dimensions[get_column_letter(qr_col)].width = QR_COL_WIDTH
    ws.column_dimensions[get_column_letter(path_col)].width = PATH_COL_WIDTH

    wb.save(output_path)

    return render_template("download.html")


# ðŸ”‘ PUBLIC QR ACCESS (VERY IMPORTANT)
@app.route("/qr/<filename>")
def serve_qr(filename):
    return send_file(os.path.join(QR_FOLDER, filename))


@app.route("/download")
def download_file():
    output_path = os.path.join(OUTPUT_FOLDER, "output_with_qr.xlsx")
    return send_file(output_path, as_attachment=True)


# ---------------- RUN LOCAL ----------------
if __name__ == "__main__":
    app.run(debug=True)
