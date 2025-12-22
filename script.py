import os
import pandas as pd
import qrcode
import tkinter as tk
from tkinter import filedialog, simpledialog, messagebox
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

# ================= CONFIG =================
QR_MM = 9
PIXELS_PER_MM = 96 / 25.4
QR_PX = int(QR_MM * PIXELS_PER_MM)
ROW_HEIGHT = 30
QR_COL_WIDTH = 18
PATH_COL_WIDTH = 95
# =========================================


def normalize(text):
    """Normalize Excel headers safely"""
    return (
        str(text)
        .replace("\xa0", " ")
        .replace("\u200b", "")
        .strip()
        .lower()
    )


def generate_qr(data, path):
    qr = qrcode.QRCode(
        version=None,
        error_correction=qrcode.constants.ERROR_CORRECT_M,
        box_size=10,
        border=1,
    )
    qr.add_data(data)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    img.save(path)


def build_qr_text(row, sheet_cols, requested_cols):
    """
    FINAL LOGIC:
    - Column has value → value
    - Column empty/missing → |
    - Separator → SPACE
    - NO '|' between two valid values
    """
    parts = []

    for col in requested_cols:
        if col not in sheet_cols:
            parts.append("|")
        else:
            val = row[col]
            if pd.isna(val) or str(val).strip() == "":
                parts.append("|")
            else:
                parts.append(str(val).strip())

    return " ".join(parts)


def main():
    # ---------- UI ----------
    root = tk.Tk()
    root.withdraw()

    excel_file = filedialog.askopenfilename(
        title="Select Excel file",
        filetypes=[("Excel files", "*.xlsx")]
    )

    if not excel_file:
        messagebox.showerror("Error", "No Excel file selected")
        return

    cols_input = simpledialog.askstring(
        "QR Columns",
        "Enter column names (comma separated):"
    )

    if not cols_input:
        messagebox.showerror("Error", "No columns entered")
        return

    requested_cols = [normalize(c) for c in cols_input.split(",")]

    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    qr_folder = os.path.join(BASE_DIR, "qr_codes")
    os.makedirs(qr_folder, exist_ok=True)

    output_file = os.path.join(BASE_DIR, "output_with_qr.xlsx")

    xls = pd.ExcelFile(excel_file)

    # ---------- WRITE ALL SHEETS ----------
    with pd.ExcelWriter(output_file, engine="openpyxl", mode="w") as writer:

        for sheet_name in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet_name)

            original_headers = list(df.columns)
            df.columns = [normalize(c) for c in df.columns]

            # Remove time from datetime columns
            for col in df.columns:
                if pd.api.types.is_datetime64_any_dtype(df[col]):
                    df[col] = df[col].dt.strftime("%Y-%m-%d")

            # Add QR columns
            df["qr_image_path"] = ""
            df["qr"] = ""

            for i, row in df.iterrows():
                qr_text = build_qr_text(row, df.columns, requested_cols)
                img_name = f"{sheet_name}_qr_{i+1}.png"
                img_path = os.path.join(qr_folder, img_name)

                generate_qr(qr_text, img_path)
                df.at[i, "qr_image_path"] = os.path.abspath(img_path)

            # Ensure column order (QR LAST)
            base_cols = [c for c in df.columns if c not in ["qr_image_path", "qr"]]
            df = df[base_cols + ["qr_image_path", "qr"]]

            # Restore readable headers
            final_headers = []
            for c in df.columns:
                restored = False
                for orig in original_headers:
                    if normalize(orig) == c:
                        final_headers.append(orig)
                        restored = True
                        break
                if not restored:
                    final_headers.append("QR_Image_Path" if c == "qr_image_path" else "QR")

            df.columns = final_headers
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    # ---------- INSERT QR IMAGES ----------
    wb = load_workbook(output_file)

    for ws in wb.worksheets:
        headers = {normalize(c.value): i + 1 for i, c in enumerate(ws[1])}
        qr_col = headers["qr"]
        path_col = headers["qr_image_path"]

        for r in range(2, ws.max_row + 1):
            img_path = ws.cell(r, path_col).value
            if not img_path or not os.path.exists(img_path):
                continue

            img = Image(img_path)
            img.width = QR_PX
            img.height = QR_PX

            cell = ws.cell(r, qr_col)
            ws.add_image(img, cell.coordinate)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[r].height = ROW_HEIGHT

        ws.column_dimensions[get_column_letter(qr_col)].width = QR_COL_WIDTH
        ws.column_dimensions[get_column_letter(path_col)].width = PATH_COL_WIDTH

    wb.save(output_file)

    messagebox.showinfo(
        "Success",
        "QR generation completed successfully!\n\n"
        "✔ All sheets processed\n"
        "✔ Original columns preserved\n"
        "✔ Correct empty-column logic\n"
        "✔ QR column LAST (image only)\n\n"
        f"Output file:\n{output_file}"
    )


if __name__ == "__main__":
    main()
